import calendar
from calendar import monthrange

from datetime import datetime
from datetime import date

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib import colors
from matplotlib.patches import Patch

from openpyxl.styles.fills import PatternFill
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import requests

from .base_requests import BaseRequests


class TableCreator(BaseRequests):
    weekdays = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    type_dict = {
        "INGOL_REST": {
            "cellValue": "R",
            "cellColor": "808080",
            "description": "Restmüll",
        },
        "INGOL_BIO": {
            "cellValue": "B",
            "cellColor": "33cc00",
            "description": "Biomüll",
        },
        "INGOL_GELB": {
            "cellValue": "G",
            "cellColor": "FFDD00",
            "description": "gelber Sack",
        },
        "INGOL_PAP": {
            "cellValue": "P",
            "cellColor": "00CCFF",
            "description": "Papiermüll",
        },
    }

    def __init__(self, user_input):
        self.selected_year = user_input.year
        if self.selected_year < date.today().year:
            self.selected_year = date.today().year
        self.end_point = user_input.location.end_point
        self.city_id = user_input.location.city_id
        self.area_id = user_input.location.area_id

        self._title = f"Müllabfuhr {str(self.selected_year)}"
        self._file_name = f"garbageCalendar{str(self.selected_year)}"
        self._address = (
            f"{user_input.street} {user_input.house_number}\n{user_input.city}"
        )
        self._num_months = 12
        self._num_days = 31

    def create_pdf(self):
        empty_cell_color = "#ffffff"

        plt.figure(
            linewidth=1,
            tight_layout={"pad": 1.5},
        )
        _ax = plt.gca()
        _ax.axis("off")

        legend_elements = []
        for _, info in self.type_dict.items():
            legend_elements.append(
                Patch(facecolor=f"#{info['cellColor']}", label=info["description"])
            )
        _ax.legend(
            handles=legend_elements,
            loc="lower left",
            bbox_to_anchor=(-0.1, 1),
            prop={"size": 5, "weight": "light"},
        )

        plt.box(on=None)
        # Add _title
        plt.suptitle(self._title)
        # Add footer
        plt.figtext(
            0.95,
            0.05,
            self._address,
            horizontalalignment="right",
            size=6,
            weight="light",
        )

        # Add a table at the bottom of the axes
        calendar_tab = plt.table(
            cellText=[[""] * (3 * self._num_months)] * self._num_days,
            rowLabels=range(1, self._num_days + 1),
            rowLoc="center",
            colLabels=[""] * (3 * self._num_months),
            loc="upper center",
            colWidths=[0.03] * (3 * self._num_months),
        )
        calendar_tab.auto_set_font_size(False)
        calendar_tab.set_fontsize(6)

        # write header
        ypos_header = 0
        for xpos in range(0, self._num_months):
            month = xpos + 1
            xpos = xpos * 3
            cell = calendar_tab[ypos_header, xpos + 2]
            cell.visible_edges = "B"
            cell = calendar_tab[ypos_header, xpos]
            cell.visible_edges = "B"
            cell = calendar_tab[ypos_header, xpos + 1]
            cell.visible_edges = "B"
            month_name = calendar.month_name[month]
            calendar_tab[ypos_header, xpos + 2].set_text_props(ha="right")
            calendar_tab[ypos_header, xpos + 2].get_text().set_text(month_name)

        # fill remaining cells:
        for xpos in range(0, self._num_months):
            month = xpos + 1
            xpos = xpos * 3

            max_days = monthrange(self.selected_year, month)
            # get calendar data
            response = requests.get(
                self.get_url_monthly_dates(self.selected_year, month)
            )
            data = response.json()["dates"]

            for day in range(1, self._num_days + 1):
                if day < max_days[1] + 1:
                    running_date = date(self.selected_year, month, day)
                    # write week day:
                    week_day = self.weekdays[running_date.weekday()]
                    cell = calendar_tab[day, xpos]
                    cell.get_text().set_text(week_day)
                    cell.set_text_props(ha="left")
                    if running_date.weekday() == 5 or running_date.weekday() == 6:
                        cell.set_text_props(weight="bold")

                    if len(data) == 0:
                        continue

                    # write all garbage info for the day:
                    for date_info in data:
                        if (
                            datetime.strptime(date_info["day"], "%Y-%m-%d").date()
                            == running_date
                        ):
                            garbage_type = date_info["trash_name"]

                            # format cell:
                            xcounter = 1
                            cell = calendar_tab[day, xpos + xcounter]
                            while (
                                xcounter < 2
                                and str(colors.to_hex(cell.get_facecolor()))
                                != empty_cell_color
                            ):
                                xcounter += 1
                                cell = calendar_tab[day, xpos + xcounter]
                            cell_info = self.type_dict.get(garbage_type)
                            cell.get_text().set_text(cell_info.get("cellValue"))
                            cell.set_text_props(ha="center")
                            cell.set_facecolor(f'#{cell_info.get("cellColor")}')
                            cell.set_edgecolor("none")

                # format cells
                cell2 = calendar_tab[day, xpos + 2]
                cell2.visible_edges = "TBR"
                cell = calendar_tab[day, xpos + 1]
                # no color in cell1
                if str(colors.to_hex(cell.get_facecolor())) == empty_cell_color:
                    cell.visible_edges = "TB"
                else:
                    # color in cell1, but not in cell2 -> merge
                    if str(colors.to_hex(cell2.get_facecolor())) == empty_cell_color:
                        cell2.set_facecolor(str(colors.to_hex(cell.get_facecolor())))
                        cell2.set_edgecolor("none")
                        cell.set_text_props(ha="right")
                cell = calendar_tab[day, xpos]
                cell.visible_edges = "LTB"

        with PdfPages(self._file_name + ".pdf") as export_pdf:
            plt.draw()
            export_pdf.savefig()
            plt.close()

    def create_excel(self):
        _wb = openpyxl.Workbook()
        _ws = _wb.active
        _ws.title = "calendar"

        empty_cell_color = "00000000"

        yoffset = 2
        xoffset = 1
        rowspan = 2
        # write heading:
        _ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        cell = _ws.cell(row=1, column=1)
        cell.value = self._title
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=True, size=32)
        _ws.row_dimensions[1].height = 40

        # border definition:
        border = Side(border_style="hair", color="000000")
        full_border = Border(top=border, left=border, right=border, bottom=border)
        top_bottom_left_border = Border(top=border, left=border, bottom=border)
        top_bottom_right_border = Border(top=border, right=border, bottom=border)
        top_bottom_border = Border(top=border, bottom=border)
        top_border = Border(top=border)

        # write numbers at border:
        _ws.column_dimensions[get_column_letter(xoffset)].width = 4
        # first empty cell:
        cell = _ws.cell(row=yoffset, column=xoffset)
        cell.border = full_border
        # number cells:
        for num in range(1, self._num_days + 1):
            cell = _ws.cell(row=num + yoffset, column=xoffset)
            cell.value = num
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.border = full_border

        # for loop over dict/array
        # write legend:
        row = 33 + yoffset
        col = xoffset + 1
        for garbage_type in self.type_dict:
            cell_info = self.type_dict.get(garbage_type)
            cell = _ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center")
            cell.value = cell_info.get("cellValue")
            color_fill = PatternFill(
                fgColor=cell_info.get("cellColor"), fill_type="solid"
            )
            cell.fill = color_fill
            cell = _ws.cell(row=row, column=col + 1)
            cell.value = ": " + cell_info.get("description")
            col = col + rowspan + 1

        for month in range(1, self._num_months + 1):
            # write month names on top:
            xpos = month * rowspan + month - rowspan + xoffset
            cell = _ws.cell(row=yoffset, column=xpos)
            cell.value = calendar.month_name[month]
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.border = full_border
            _ws.merge_cells(
                start_row=yoffset,
                start_column=xpos,
                end_row=yoffset,
                end_column=xpos + rowspan,
            )
            max_days = monthrange(self.selected_year, month)
            # format weekday column:
            _ws.column_dimensions[get_column_letter(xpos)].width = 4
            for col_counter in range(1, rowspan + 1):
                _ws.column_dimensions[get_column_letter(xpos + col_counter)].width = 6

            # get calendar data
            response = requests.get(
                self.get_url_monthly_dates(self.selected_year, month)
            )
            data = response.json()["dates"]

            # loop through days of month:
            for day in range(1, max_days[1] + 1):
                running_date = date(self.selected_year, month, day)
                ypos = yoffset + day
                # write weekday:
                cell = _ws.cell(row=ypos, column=xpos)
                cell.value = self.weekdays[running_date.weekday()]
                # print "Sa" and "So" bold:
                if running_date.weekday() == 5 or running_date.weekday() == 6:
                    cell.font = Font(bold=True)
                # put borders
                cell.border = full_border
                for col_counter in range(1, rowspan + 1):
                    cell = _ws.cell(row=ypos, column=xpos + col_counter)
                    if col_counter == 1:
                        cell.border = top_bottom_left_border
                    elif col_counter == rowspan:
                        cell.border = top_bottom_right_border
                    else:
                        cell.border = top_bottom_border

                if len(data) == 0:
                    continue

                # write all garbage info for the day:
                for date_info in data:
                    if (
                        datetime.strptime(date_info["day"], "%Y-%m-%d").date()
                        == running_date
                    ):
                        garbage_type = date_info["trash_name"]

                        # format cell:
                        xcounter = 1
                        cell = _ws.cell(row=ypos, column=xpos + xcounter)
                        while (
                            xcounter < rowspan
                            and cell.fill.start_color.index != empty_cell_color
                        ):
                            xcounter += 1
                            cell = _ws.cell(row=ypos, column=xpos + xcounter)
                        cell.alignment = Alignment(horizontal="center")
                        cell_info = self.type_dict.get(garbage_type)
                        cell.value = cell_info.get("cellValue")
                        color_fill = PatternFill(
                            # fgColor=date_info["color"],
                            fgColor=cell_info.get("cellColor"),
                            fill_type="solid",
                        )
                        cell.fill = color_fill

                # merge cells if only one entry:
                cell = _ws.cell(row=ypos, column=xpos + rowspan)
                if cell.fill.start_color.index == empty_cell_color:
                    _ws.merge_cells(
                        start_row=ypos,
                        start_column=xpos + 1,
                        end_row=ypos,
                        end_column=xpos + rowspan,
                    )

            # put border in last cell of month:
            for col_counter in range(0, rowspan + 1):
                cell = _ws.cell(row=yoffset + 32, column=xpos + col_counter)
                cell.border = top_border

        # define print area
        _ws.print_area = "A1:AK35"
        # set to page landscape mode
        _ws.set_printer_settings(
            paper_size=openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_A4,
            orientation=openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_LANDSCAPE,
        )
        # define scaling when printing
        _ws.sheet_properties.pageSetUpPr.fitToPage = True
        _ws.page_setup.fitToHeight = False

        _wb.save(self._file_name + ".xls")
