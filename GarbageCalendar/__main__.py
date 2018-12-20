#!/usr/bin/python
import requests

from openpyxl.styles.fills import PatternFill
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from datetime import datetime
from datetime import date

import calendar
from calendar import monthrange

import tkinter as tk 
from tkinter import ttk # additional tkinter elements like combo boxes

import sys

## open gui (pipe response from api through) and ask for street and housenumber.
## name is street name, houseNumbers has array of house numbers [number, id, comment]
response = requests.get("https://ingol.jumomind.com/mmapp/api.php?r=streets&city_id=87")
data = response.json()
area_id = ""
selected_street = ""
    
## interact with user:
query_window = tk.Tk()  
query_window.title("Additional information required")

tk.Label( query_window, text = 'Street: '    , anchor = 'e' ).grid( row = 1, column = 1 )
tk.Label( query_window, text = 'House Number: '    , anchor = 'e' ).grid( row = 2, column = 1 )

street_chooser = ttk.Combobox(query_window, width = 27, textvariable = tk.StringVar()) 
result = []
for street in data:
    result.append(street['name'])
street_chooser['values'] = tuple(result)
street_chooser.grid(row = 1, column = 2) 
 
number_chooser = ttk.Combobox(query_window, width = 27, textvariable = tk.StringVar()) 
number_chooser.grid(row = 2, column = 2) 
 
def setHouseNumbers(event):
    global selected_street
    street_name = street_chooser.get()
    
    result = []
    for street in data:
        if street['name'] == street_name:
            selected_street = street
            for house_number in selected_street['houseNumbers']:
                result.append(house_number[0])
            number_chooser['values'] = tuple(result)
            break
        number_chooser.bind("<<ComboboxSelected>>", startGeneration)

def handle_street_keyrelease(event):
    typed = street_chooser.get()
    result = []
    for street in data:
        if street['name'].startswith(typed):
            result.append(street['name'])
    street_chooser['values'] = tuple(result)
    street_chooser.event_generate('<Down>')
    

def startGeneration(event):
    global area_id 
    for number in selected_street['houseNumbers']:
        selected_house_number = number_chooser.get()
        if number[0].__eq__(selected_house_number):
            area_id = number[1]
            break
    query_window.destroy()

street_chooser.bind("<<ComboboxSelected>>", setHouseNumbers)
street_chooser.bind('<KeyRelease>', handle_street_keyrelease)

query_window.mainloop() 

if area_id.__eq__(""):
    sys.exit("the data you entered is not complete")
## parse params from input
    
params = {
        "idx": "termins",
        "city_id": "87",
        "area_id": area_id,
        "ws": "3"
        }

response = requests.get("https://ingol.jumomind.com/webservice.php", params)

data = response.json()[0]['_data']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "calendar"

thisYear = date.today().year

yOffset = 2
xOffset = 1
weekdays = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
rowspan = 2
## dict with names returned from request as key, containing cellColor and cellValue
type_dict = {
    "INGOL_REST": {
                    "cellValue": "R",
                    "cellColor": '808080',
                    "description": "Restmüll"
                    },
    "INGOL_BIO": {
                    "cellValue": "B",            
                    "cellColor": '33cc00',
                    "description": "Biomüll"
        },
    "INGOL_GELB":{
                    "cellValue": "G",              
                    "cellColor": 'FFDD00',
                    "description": "gelber Sack"
        },
    "INGOL_PAP":{
                    "cellValue": "P",             
                    "cellColor": '00CCFF',
                    "description": "Papiermüll"
        }
    }

# write heading:
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
cell = ws.cell(row=1, column=1)
cell.value = "Müllabfuhr " + str(thisYear)
cell.alignment = Alignment(horizontal='left')
cell.font = Font(bold=True, size=32)
ws.row_dimensions[1].height = 40

# border definition:
border = Side(border_style="hair", color="000000")
full_border = Border(top=border, left=border, right=border, bottom=border)
top_bottom_left_border = Border(top=border, left=border, bottom=border)
top_bottom_right_border = Border(top=border, right=border, bottom=border)
top_bottom_border = Border(top=border, bottom=border)
top_border = Border(top=border)

# write numbers at border:
ws.column_dimensions[get_column_letter(xOffset)].width = 4
# first empty cell:
cell=ws.cell(row=yOffset, column=xOffset)
cell.border = full_border
# number cells:
for num in range(1,32):
    cell=ws.cell(row=num+yOffset, column=xOffset)
    cell.value = num
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    cell.border = full_border
    

# for loop over dict/array
# write legend:
row = 33+yOffset
col = xOffset+1
for garbage_type in type_dict:
    cellInfo = type_dict.get(garbage_type)
    cell = ws.cell(row=row, column=col)
    cell.alignment = Alignment(horizontal='center')
    cell.value = cellInfo.get("cellValue")
    colorFill = PatternFill(fgColor=cellInfo.get("cellColor"), fill_type='solid')
    cell.fill = colorFill
    cell = ws.cell(row=row, column=col+1)
    cell.value = ": " + cellInfo.get("description")
    col = col + rowspan + 1



counter = 0

for month in range(1, 13):
    # write month names on top:
    xPos = month*rowspan+month-rowspan+xOffset
    cell = ws.cell(row=yOffset, column=xPos)
    cell.value = calendar.month_name[month]
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    cell.border = full_border
    ws.merge_cells(start_row=yOffset, start_column=xPos, end_row=yOffset, end_column=xPos+rowspan)
    maxDays = monthrange(thisYear, month)
    # format weekday column:
    ws.column_dimensions[get_column_letter(xPos)].width = 4
    for colCounter in range(1,rowspan+1):
        ws.column_dimensions[get_column_letter(xPos+colCounter)].width = 6
        

    # loop through days of month:
    for day in range(1,maxDays[1]+1):
        runningDate = date(thisYear, month, day)
        yPos = yOffset+day
        # write weekday:
        cell = ws.cell(row=yPos, column=xPos)
        cell.value = weekdays[runningDate.weekday()]
        # print "Sa" and "So" bold:
        if(runningDate.weekday() == 5 or runningDate.weekday() == 6):
            cell.font = Font(bold=True)
        # put bprders
        cell.border = full_border
        for colCounter in range(1,rowspan+1):
            cell = ws.cell(row=yPos, column=xPos+colCounter)
            if(colCounter == 1):
                cell.border = top_bottom_left_border
            elif(colCounter==rowspan):
                cell.border = top_bottom_right_border
            else:
                cell.border = top_bottom_border

        # write all garbage info for the day:
        if(datetime.strptime(data[counter]['cal_date'], "%Y-%m-%d").date() == runningDate):
            xcounter = 1
            
            while(datetime.strptime(data[counter]['cal_date'], "%Y-%m-%d").date() == runningDate):
                garbage_type = data[counter]['cal_garbage_type']

                # format cell:
                cell = ws.cell(row=yPos, column=xPos+xcounter)
                cell.alignment = Alignment(horizontal='center')
                cellInfo = type_dict.get(garbage_type)
                cell.value = cellInfo.get("cellValue")
                colorFill = PatternFill(fgColor=cellInfo.get("cellColor"), fill_type='solid')
                cell.fill = colorFill
                counter += 1
                xcounter += 1
                
            # merge cells if only one entry:
            if(xcounter == xOffset+1):
                ws.merge_cells(start_row=yPos, start_column=xPos+1, end_row=yPos, end_column=xPos+2)
                
    
    # put border in last cell of month:
    for colCounter in range(0,rowspan+1):
        cell=ws.cell(row=yOffset+32, column=xPos+colCounter)
        cell.border = top_border
                

wb.save("garbageCalendar"+str(thisYear)+".xls")
print('finished')




